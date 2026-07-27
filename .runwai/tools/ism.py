#!/usr/bin/env python3
"""ISM snapshot and verification for runwAI.

Two subcommands:

  snapshot  Extract ISM controls from an ASD template (.xlsx) into
            controls/ism-snapshot.json. Requires openpyxl and the template.
            Run once per ISM release.

  verify    Check every ISM ID claimed in controls/registry.yaml against the
            snapshot, and compare each claim's recorded description against the
            authoritative text. Offline, no network, no LLM: same snapshot plus
            same registry always yields the same verdict.

  index     Join the snapshot, the semantic tags in controls/ism-tags.yaml and the
            rationale prose in controls/ism-source.txt into controls/ism-index.json:
            one machine-readable record per control, tagged and surfaced. Generated,
            so `index --check` fails when the committed file is stale.

The split is deliberate. Extraction needs a source document that is not
redistributable in its original form; verification must run in CI on every
commit. Snapshotting once and verifying offline makes the check reproducible.

ISM control text in the snapshot is (c) Commonwealth of Australia, released
under CC BY 4.0. See controls/ism-snapshot.json for full attribution.

Exit codes:
    0  verification passed
    1  verification failed
    2  could not run (missing dependency, template or snapshot)
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover
    print("FATAL: PyYAML required. pip install pyyaml==6.0.3", file=sys.stderr)
    raise SystemExit(2)


SNAPSHOT_REL = Path("controls") / "ism-snapshot.json"

ATTRIBUTION = {
    "source": "Australian Signals Directorate, Information security manual",
    "release": "June 2026",
    "derived_from": (
        "System security plan annex template (June 2026), 'Controls - June 2026' "
        "worksheet, published alongside the ISM on the ASD website."
    ),
    "copyright": "(c) Commonwealth of Australia",
    "license": "CC-BY-4.0",
    "license_url": "https://creativecommons.org/licenses/by/4.0/legalcode.en",
    "note": (
        "Control identifiers and descriptions are reproduced under CC BY 4.0 for "
        "verification purposes. The Commonwealth Coat of Arms is excluded from that "
        "licence and is not reproduced here. This snapshot is not an ASD publication "
        "and carries no ASD endorsement."
    ),
}

# Worksheet column positions in the ASD template. Asserted against the header
# row at extraction time so a template reshuffle fails loudly instead of
# silently importing the wrong column.
COLUMNS = {
    "guideline": (0, "Guideline"),
    "section": (1, "Section"),
    "topic": (2, "Topic"),
    "identifier": (3, "Identifier"),
    "revision": (4, "Revision"),
    "updated": (5, "Updated"),
    "description": (14, "Description"),
}
APPLICABILITY = {"NC": 6, "OS": 7, "P": 8, "S": 9, "TS": 10}
MATURITY = {"ML1": 11, "ML2": 12, "ML3": 13}

SHEET_NAME = "Controls - June 2026"


def cmd_snapshot(args: argparse.Namespace) -> int:
    try:
        import openpyxl
    except ImportError:
        print("FATAL: openpyxl required for snapshotting. pip install openpyxl", file=sys.stderr)
        return 2

    template = Path(args.template)
    if not template.is_file():
        print(f"FATAL: template not found: {template}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(template, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(
            f"FATAL: worksheet {SHEET_NAME!r} not found. Sheets: {wb.sheetnames}",
            file=sys.stderr,
        )
        return 2
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)
    for key, (idx, expected) in COLUMNS.items():
        actual = str(header[idx]).strip() if header[idx] else ""
        if actual != expected:
            print(
                f"FATAL: column {idx} is {actual!r}, expected {expected!r}. "
                "The template layout changed; update COLUMNS before snapshotting.",
                file=sys.stderr,
            )
            return 2

    controls: dict[str, dict] = {}
    for row in rows:
        ident = row[COLUMNS["identifier"][0]]
        if not ident:
            continue
        ident = str(ident).strip()
        controls[ident] = {
            "guideline": row[COLUMNS["guideline"][0]],
            "section": row[COLUMNS["section"][0]],
            "topic": row[COLUMNS["topic"][0]],
            "revision": row[COLUMNS["revision"][0]],
            "updated": row[COLUMNS["updated"][0]],
            "description": " ".join(str(row[COLUMNS["description"][0]] or "").split()),
            "applicable": [k for k, i in APPLICABILITY.items() if row[i] == "Yes"],
            "essential_eight": [k for k, i in MATURITY.items() if row[i] == "Yes"],
        }
    wb.close()

    payload = {
        "attribution": ATTRIBUTION,
        "control_count": len(controls),
        "template_sha256": hashlib.sha256(template.read_bytes()).hexdigest(),
        "controls": dict(sorted(controls.items())),
    }

    out = Path(args.repo_root) / SNAPSHOT_REL
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=1, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"wrote {out} ({len(controls)} controls)")
    print(f"template sha256: {payload['template_sha256']}")
    return 0


def normalise(text: str) -> str:
    return " ".join(str(text or "").split()).lower().rstrip(".")


def cmd_verify(args: argparse.Namespace) -> int:
    root = Path(args.repo_root).resolve()

    snapshot_path = root / SNAPSHOT_REL
    if not snapshot_path.is_file():
        print(
            f"FATAL: {SNAPSHOT_REL} not found. Run: python3 .runwai/tools/ism.py snapshot "
            "--template <ASD template.xlsx>",
            file=sys.stderr,
        )
        return 2
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    ism = snapshot["controls"]

    registry_path = root / "controls" / "registry.yaml"
    registry = yaml.safe_load(registry_path.read_text(encoding="utf-8"))

    errors: list[str] = []
    warnings: list[str] = []
    claims = 0

    for ctl in registry.get("controls", []):
        rwa = ctl.get("id")
        recorded = ctl.get("ism_text", {}) or {}

        for ident in ctl.get("ism_ids", []):
            claims += 1
            if ident not in ism:
                errors.append(
                    f"{rwa} claims {ident}, which does not exist in the "
                    f"{snapshot['attribution']['release']} ISM"
                )
                continue

            authoritative = ism[ident]["description"]

            if ident not in recorded:
                warnings.append(
                    f"{rwa} claims {ident} but records no ism_text for it, so drift "
                    "in the authoritative wording cannot be detected"
                )
                continue

            if normalise(recorded[ident]) != normalise(authoritative):
                errors.append(
                    f"{rwa}/{ident} recorded text does not match the ISM.\n"
                    f"    recorded: {recorded[ident]}\n"
                    f"    ISM     : {authoritative}"
                )

        for ident in recorded:
            if ident not in ctl.get("ism_ids", []):
                errors.append(f"{rwa} records ism_text for {ident} which it does not claim")

    release = registry.get("ism_release", {})
    if release.get("verification_status") == "verified":
        if release.get("control_count") != snapshot["control_count"]:
            errors.append(
                f"registry ism_release.control_count is {release.get('control_count')}, "
                f"snapshot holds {snapshot['control_count']}"
            )

    for w in warnings:
        print(f"WARN  {w}")
    for e in errors:
        print(f"ERROR {e}")

    print(
        f"\nrunwai-ism-verify: {claims} ISM claims checked against "
        f"{snapshot['control_count']} controls "
        f"({snapshot['attribution']['release']}), "
        f"{len(errors)} errors, {len(warnings)} warnings"
    )
    if errors:
        print("FAILED")
        return 1
    if args.strict and warnings:
        print("FAILED (strict: warnings are errors)")
        return 1
    print("PASSED")
    return 0


# "Control: ISM-0407; Revision: 6; Updated: Jun-25; Applicable: ...; Essential 8: N/A"
CONTROL_MARK = re.compile(r"^Control:\s*(ISM-\d{4});", re.M)

# Running header repeated on every page of the published PDF-to-text conversion.
PAGE_FURNITURE = re.compile(r"^\s*Information security manual\s+\S*\s*$", re.M)


def load_json(path: Path, errors: list[str]) -> dict | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        errors.append(f"{path.name}: {exc}")
        return None


def load_yaml_file(path: Path, errors: list[str]) -> dict | None:
    try:
        return yaml.safe_load(path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError) as exc:
        errors.append(f"{path.name}: {exc}")
        return None


def flatten(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def extract_rationale(source: str, controls: dict) -> dict[str, str]:
    """Pull each topic's explanatory prose out of the published ISM text.

    The snapshot keeps only the control sentence. The published text carries, before each
    group of controls, a sub-heading and a paragraph explaining *why* the controls exist —
    which is the best available signal for what a control is about, and the only way to
    explain one in the ISM's own words rather than paraphrasing it.

    Located by subtraction rather than by pattern: everything between the end of the
    previous control's known text and the next control marker is that control's preamble.
    The preceding text is known exactly, from the snapshot, so this needs no guess about
    where a control's own wording stops.

    Rationale attaches to a topic, not to a control — 450 topics, and the published text
    carries one block per topic. Controls in a topic share it. A topic whose block cannot
    be located is absent from the result rather than approximated.
    """
    marks = [(m.group(1), m.start(), m.end()) for m in CONTROL_MARK.finditer(source)]
    out: dict[str, str] = {}
    for i, (cid, start, _) in enumerate(marks):
        if i == 0 or cid not in controls:
            continue
        prev_id, _, prev_end = marks[i - 1]
        if prev_id not in controls:
            continue
        prev_text = flatten(controls[prev_id]["description"])
        between = flatten(PAGE_FURNITURE.sub(" ", source[prev_end:start]))
        # 80 characters is enough to be unambiguous and short enough to survive the
        # whitespace and hyphenation damage of the published conversion.
        at = between.find(prev_text[:80]) if prev_text else -1
        if at < 0:
            continue
        tail = between[at + len(prev_text):].strip()
        if not tail:
            continue
        topic = controls[cid]["topic"]
        # The block opens with the sub-heading, which is the topic we already carry.
        if tail.startswith(topic):
            tail = tail[len(topic):].strip()
        if tail and topic not in out:
            out[topic] = tail
    return out


def build_index(root: Path) -> tuple[dict | None, list[str]]:
    """Join snapshot + tags + rationale. Returns (index, errors)."""
    errors: list[str] = []

    snapshot = load_json(root / "controls" / "ism-snapshot.json", errors)
    tagfile = load_yaml_file(root / "controls" / "ism-tags.yaml", errors)
    if snapshot is None or tagfile is None:
        return None, errors

    controls = snapshot.get("controls", {})
    vocabulary = set(tagfile.get("vocabulary", {}))
    surfaces = set(tagfile.get("surfaces", []))
    sections = tagfile.get("sections", {})
    overrides = tagfile.get("overrides", {}) or {}

    def check(where: str, entry: dict) -> None:
        for tag in entry.get("tags", []):
            if tag not in vocabulary:
                errors.append(f"{where}: tag '{tag}' is not in the vocabulary")
        got = entry.get("surfaces", [])
        if not got:
            errors.append(f"{where}: declares no surfaces")
        for s in got:
            if s not in surfaces:
                errors.append(f"{where}: surface '{s}' is not one of {sorted(surfaces)}")

    for guideline, secs in sections.items():
        for section, entry in secs.items():
            check(f"sections/{guideline}/{section}", entry)

    for cid, entry in overrides.items():
        where = f"overrides/{cid}"
        if cid not in controls:
            errors.append(f"{where}: no such control in the snapshot")
        if not str(entry.get("reason", "")).strip():
            errors.append(f"{where}: has no reason. An override without an argument is "
                          "indistinguishable from a mistake.")
        check(where, entry)

    # Every section present in the snapshot must be tagged. A new ISM release that adds one
    # fails here rather than silently producing untagged controls.
    for cid, ctl in controls.items():
        if ctl["section"] not in sections.get(ctl["guideline"], {}):
            errors.append(
                f"controls/ism-tags.yaml: no entry for '{ctl['guideline']}' / "
                f"'{ctl['section']}', which {cid} is in"
            )
            break

    if errors:
        return None, errors

    source_path = root / "controls" / "ism-source.txt"
    rationale = (
        extract_rationale(source_path.read_text(encoding="utf-8"), controls)
        if source_path.is_file()
        else {}
    )

    records = {}
    for cid in sorted(controls):
        ctl = controls[cid]
        entry = sections[ctl["guideline"]][ctl["section"]]
        origin = "section"
        if cid in overrides:
            entry = overrides[cid]
            origin = "override"
        records[cid] = {
            "guideline": ctl["guideline"],
            "section": ctl["section"],
            "topic": ctl["topic"],
            "description": ctl["description"],
            "applicable": ctl.get("applicable", []),
            "essential_eight": ctl.get("essential_eight", []),
            "surfaces": sorted(entry["surfaces"]),
            "tags": sorted(entry.get("tags", [])),
            "tagged_by": origin,
            "rationale": rationale.get(ctl["topic"]),
        }

    # Counts per surface, not a derived "automatable" number. A surface says where the
    # evidence for a control lives, never that runwAI can reach it: ISM-0260 ("all web
    # access is conducted through web proxies") is infrastructure and no repository
    # declares it. Deriving an automatable count from these would be the coverage claim
    # this repository exists to avoid making.
    per_surface = {s: 0 for s in sorted(surfaces)}
    for record in records.values():
        for s in record["surfaces"]:
            per_surface[s] += 1

    return {
        "attribution": snapshot.get("attribution"),
        "note": (
            "Generated by `python3 .runwai/tools/ism.py index`. Do not edit by hand — edit "
            "controls/ism-tags.yaml and regenerate. Control text and rationale are (c) "
            "Commonwealth of Australia under CC BY 4.0; the surfaces and tags are runwAI's "
            "judgement, Apache-2.0. A surface says where a control's evidence lives, not "
            "that runwAI can reach it. Nothing here is a compliance claim."
        ),
        "control_count": len(records),
        "controls_per_surface": per_surface,
        "rationale_coverage": sum(1 for r in records.values() if r["rationale"]),
        "controls": records,
    }, []


def cmd_index(args: argparse.Namespace) -> int:
    root = Path(args.repo_root).resolve()
    index, errors = build_index(root)
    if errors:
        for e in errors:
            print(f"ERROR {e}")
        print(f"\nrunwai-ism-index: {len(errors)} errors\nFAILED")
        return 1

    # Sorted keys and a trailing newline: byte-identical output for an unchanged tree, so
    # --check compares content rather than churn.
    text = json.dumps(index, indent=2, ensure_ascii=False, sort_keys=True) + "\n"
    path = root / "controls" / "ism-index.json"

    surface_summary = ", ".join(
        f"{n} {s}" for s, n in sorted(index["controls_per_surface"].items())
    )
    summary = (
        f"runwai-ism-index: {index['control_count']} controls ({surface_summary}), "
        f"{index['rationale_coverage']} carrying rationale"
    )

    if args.check:
        if not path.is_file():
            print(f"FAILED: {path.name} does not exist. Run: python3 .runwai/tools/ism.py index")
            return 1
        if path.read_text(encoding="utf-8") != text:
            print(
                f"FAILED: {path.name} is stale. Run `python3 .runwai/tools/ism.py index` "
                "and commit the result."
            )
            return 1
        print(summary)
        print("PASSED")
        return 0

    path.write_text(text, encoding="utf-8")
    print(summary)
    print("PASSED")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", default=".")
    sub = parser.add_subparsers(dest="cmd", required=True)

    snap = sub.add_parser("snapshot", help="extract ISM controls from an ASD template")
    snap.add_argument("--template", required=True, help="path to the ASD .xlsx template")
    snap.set_defaults(func=cmd_snapshot)

    ver = sub.add_parser("verify", help="verify registry ISM claims against the snapshot")
    ver.add_argument("--strict", action="store_true", help="treat warnings as failures")
    ver.set_defaults(func=cmd_verify)

    idx = sub.add_parser("index", help="generate controls/ism-index.json from snapshot + tags")
    idx.add_argument("--check", action="store_true", help="fail if the committed index is stale")
    idx.set_defaults(func=cmd_index)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
